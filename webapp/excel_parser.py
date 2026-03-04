"""
엑셀/CSV 파일에서 날짜별 소진액 데이터를 파싱한다.

엑셀 구조:
- Row 2: 날짜 헤더 (예: "3월 1일 소진액", "3월 2일 소진액") — 병합 셀일 수 있음
- Row 3: 컬럼 헤더 (NO., 광고주ID, 계정ID, 광고주명, 매체, 소진액, 차액)
- Row 4~: 데이터 행 (마지막은 합계 행)
- 날짜 블록이 좌→우 나란히, 빈 컬럼으로 구분

CSV 구조 (네이버 에이전트 실적 리포트, EUC-KR):
- [01] 광고주ID, [02] 계정ID, [04] 광고주명
- [44] 성과형 DA 유상 매출 → GFA
- [46] ADVoost 쇼핑 유상 매출 → AD
- [47] 유상실적TOTAL → 전체 합계, 네이버 = TOTAL - GFA - AD
"""

import re
import csv
from datetime import date
import openpyxl


DATE_HEADER_PATTERN = re.compile(r'(\d{1,2})월\s*(\d{1,2})일')


def _get_cell_value(ws, row, col):
    """병합 셀 포함하여 실제 값을 반환한다."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    # 병합 셀이면 병합 범위의 최상단-좌측 셀 값을 반환
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col):
            return ws.cell(row=merged.min_row, column=merged.min_col).value
    return None


def _to_int(value):
    """숫자로 변환 가능한 값이면 int, 아니면 0을 반환한다."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        cleaned = value.replace(',', '').replace(' ', '')
        try:
            return int(float(cleaned))
        except ValueError:
            return 0
    return 0


def parse_excel(filepath, year=None):
    """
    엑셀 파일을 파싱해 소진액 레코드 목록을 반환한다.

    Returns:
        list of dict: [
            {
                'date': date(2026, 3, 1),
                'advertiser_id': 'mianso:naver',
                'account_id': '717609',
                'advertiser_name': '명도',
                'media': '네이버',
                'amount': 6660,
            },
            ...
        ]
    """
    if year is None:
        year = date.today().year

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 첫 번째 시트 사용 (또는 "소진액체크" 시트 우선)
    ws = None
    for name in wb.sheetnames:
        if '소진액' in name:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    # ── Step 1: Row 3에서 "소진액" 컬럼 위치 수집 ──────────────────
    spend_cols = []  # [(col_index, header_text), ...]
    for col in range(1, max_col + 1):
        val = ws.cell(row=3, column=col).value
        if val and '소진액' in str(val):
            spend_cols.append(col)

    if not spend_cols:
        raise ValueError("엑셀에서 '소진액' 컬럼을 찾을 수 없습니다. (Row 3 기준)")

    # ── Step 2: 각 소진액 컬럼에 대해 날짜와 보조 컬럼 매핑 ──────
    date_blocks = []  # [{'date': date, 'spend_col': col, 'id_col': col, 'name_col': col, 'media_col': col}, ...]
    for spend_col in spend_cols:
        # Row 2에서 날짜 헤더 탐색 (spend_col 포함 블록)
        block_date = None
        # spend_col 기준으로 왼쪽 6개 컬럼 범위 내 Row 2 탐색
        search_range = range(max(1, spend_col - 7), spend_col + 1)
        for c in reversed(search_range):
            row2_val = _get_cell_value(ws, 2, c)
            if row2_val:
                m = DATE_HEADER_PATTERN.search(str(row2_val))
                if m:
                    month, day = int(m.group(1)), int(m.group(2))
                    try:
                        block_date = date(year, month, day)
                    except ValueError:
                        pass
                    break

        if block_date is None:
            continue  # 날짜를 파싱할 수 없으면 스킵

        # 보조 컬럼 위치: spend_col 기준 왼쪽으로
        # 구조: NO(-5) 광고주ID(-4) 계정ID(-3) 광고주명(-2) 매체(-1) 소진액(0)
        id_col = spend_col - 4
        account_col = spend_col - 3
        name_col = spend_col - 2
        media_col = spend_col - 1

        date_blocks.append({
            'date': block_date,
            'spend_col': spend_col,
            'id_col': id_col,
            'account_col': account_col,
            'name_col': name_col,
            'media_col': media_col,
        })

    if not date_blocks:
        raise ValueError("엑셀에서 날짜 정보를 파싱할 수 없습니다.")

    # ── Step 3: 데이터 행 순회 (Row 4부터) ──────────────────────────
    records = []
    for row in range(4, max_row + 1):
        # 빈 행 스킵
        row_values = [ws.cell(row=row, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in row_values):
            continue

        for block in date_blocks:
            id_val = ws.cell(row=row, column=block['id_col']).value
            name_val = ws.cell(row=row, column=block['name_col']).value
            media_val = ws.cell(row=row, column=block['media_col']).value
            spend_val = ws.cell(row=row, column=block['spend_col']).value
            account_val = ws.cell(row=row, column=block['account_col']).value

            # 합계 행 스킵 (광고주ID가 없거나 "합계" 텍스트 포함)
            if not id_val or (isinstance(id_val, str) and '합계' in id_val):
                continue
            if not name_val or (isinstance(name_val, str) and '합계' in name_val):
                continue
            if not media_val:
                continue

            advertiser_id = str(id_val).strip()
            advertiser_name = str(name_val).strip()
            media = str(media_val).strip()
            amount = _to_int(spend_val)
            account_id = str(account_val).strip() if account_val else ''

            # 소진액 0이면 스킵
            if amount == 0:
                continue

            records.append({
                'date': block['date'],
                'advertiser_id': advertiser_id,
                'account_id': account_id,
                'advertiser_name': advertiser_name,
                'media': media,
                'amount': amount,
            })

    return records


def preview_parse(filepath, year=None):
    """파싱 결과 요약을 반환한다 (업로드 미리보기용)."""
    records = parse_excel(filepath, year)
    return _make_preview(records)


# ── CSV 파싱 (네이버 에이전트 실적 리포트) ──────────────────────────

# CSV 컬럼 인덱스
_CSV_ADVERTISER_ID = 1
_CSV_ACCOUNT_ID = 2
_CSV_ADVERTISER_NAME = 4
_CSV_GFA = 44          # 성과형 DA 유상 매출
_CSV_AD = 46           # ADVoost 쇼핑 유상 매출
_CSV_TOTAL = 47        # 유상실적TOTAL


def parse_csv(filepath, target_date):
    """
    네이버 에이전트 실적 CSV를 파싱해 소진액 레코드 목록을 반환한다.

    Args:
        filepath: CSV 파일 경로
        target_date: date 객체 (CSV에 날짜 정보가 없으므로 직접 지정)

    Returns:
        list of dict (parse_excel과 동일한 형식)
    """
    records = []

    # EUC-KR 시도 → 실패하면 UTF-8
    for encoding in ('euc-kr', 'cp949', 'utf-8'):
        try:
            with open(filepath, encoding=encoding, errors='strict') as f:
                f.read(100)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        encoding = 'utf-8'

    with open(filepath, encoding=encoding, errors='replace') as f:
        reader = csv.reader(f)
        rows = list(reader)

    if len(rows) < 2:
        raise ValueError("CSV 파일에 데이터가 없습니다.")

    # 헤더 확인 (최소한 광고주ID 컬럼 존재 여부)
    header = rows[0]
    if len(header) < _CSV_TOTAL + 1:
        raise ValueError(f"CSV 컬럼 수가 부족합니다. (필요: {_CSV_TOTAL + 1}개, 실제: {len(header)}개)")

    for row in rows[1:]:
        if len(row) < _CSV_TOTAL + 1:
            continue

        advertiser_id = row[_CSV_ADVERTISER_ID].strip()
        account_id = row[_CSV_ACCOUNT_ID].strip()
        advertiser_name = row[_CSV_ADVERTISER_NAME].strip()

        # 필수 값 없으면 스킵
        if not advertiser_id or not advertiser_name:
            continue
        # '-' 이름은 실제 이름이 없는 경우 — 광고주ID를 이름으로 사용
        if advertiser_name == '-':
            advertiser_name = advertiser_id

        total = _to_int(row[_CSV_TOTAL])
        gfa = _to_int(row[_CSV_GFA])
        ad = _to_int(row[_CSV_AD])
        naver = total - gfa - ad

        # 전체 0이면 스킵
        if total == 0:
            continue

        # 매체별 레코드 생성 (금액 > 0인 경우만)
        if naver > 0:
            records.append({
                'date': target_date,
                'advertiser_id': advertiser_id,
                'account_id': account_id,
                'advertiser_name': advertiser_name,
                'media': '네이버',
                'amount': naver,
            })

        if gfa > 0:
            records.append({
                'date': target_date,
                'advertiser_id': advertiser_id,
                'account_id': account_id,
                'advertiser_name': advertiser_name,
                'media': 'GFA',
                'amount': gfa,
            })

        if ad > 0:
            records.append({
                'date': target_date,
                'advertiser_id': advertiser_id,
                'account_id': account_id,
                'advertiser_name': advertiser_name,
                'media': 'AD',
                'amount': ad,
            })

    return records


def preview_parse_csv(filepath, target_date):
    """CSV 파싱 결과 요약을 반환한다."""
    records = parse_csv(filepath, target_date)
    return _make_preview(records)


# ── 공통 미리보기 헬퍼 ──────────────────────────────────────────────

def _make_preview(records):
    """레코드 목록에서 미리보기 요약을 생성한다."""
    dates = sorted(set(r['date'] for r in records))
    advertisers = sorted(set(r['advertiser_name'] for r in records))
    medias = sorted(set(r['media'] for r in records))
    total_amount = sum(r['amount'] for r in records)
    return {
        'records': records,
        'dates': dates,
        'advertisers': advertisers,
        'medias': medias,
        'total_amount': total_amount,
        'record_count': len(records),
    }
